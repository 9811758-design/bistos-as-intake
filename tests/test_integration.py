import hashlib
import zipfile
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from service_validation.domain import ValidationError
from service_validation.issuance import (
    append_issuance_log,
    load_issued_counts,
    record_key,
    write_status_report,
)
from service_validation.report_writer import PASS_COUNTS, generate_workbook
from service_validation.service import create_reports
from service_validation.workbook import RejectedRow, SourceRows, load_overrides, read_source

SOURCE = Path(r"C:\Users\User\Downloads\엑셀시트.xlsx")
TEMPLATE = Path(r"C:\Users\User\Downloads\서비스검증결과서 양식.xlsx")


def test_generate_sample_workbook_when_real_files_available(tmp_path: Path) -> None:
    source_hash = hashlib.sha256(SOURCE.read_bytes()).digest()
    template_hash = hashlib.sha256(TEMPLATE.read_bytes()).digest()
    rows = read_source(SOURCE)
    record = next(row for row in rows.records if row.service_number == "DS26070601")
    output = generate_workbook(TEMPLATE, tmp_path, record)
    with zipfile.ZipFile(output) as package:
        names = set(package.namelist())
        assert any(name.startswith("xl/drawings/drawing") for name in names)
        assert any(name.startswith("xl/media/image") for name in names)
        assert b"<drawing" in package.read("xl/worksheets/sheet1.xml")
    workbook = load_workbook(output)
    assert workbook.sheetnames == ["BT380"]
    sheet = workbook["BT380"]
    assert sheet["BO1"].value == "DS26070601"
    assert sheet["O4"].value == "2026 년 07 월 06 일"
    assert sheet["O5"].value == "■대리점      □병원      □개인고객"
    assert sheet["BE5"].value == "국제메디칼(미즈피아병원)"
    assert sheet["O9"].value == "화면이나오지 않음"
    assert sheet["O12"].value == "Display cable 교체"
    assert sheet["O12"].font.sz == 10
    assert sheet["O13"].value == "2026 년 07 월 06 일"
    pass_cells = [
        cell
        for row in sheet.iter_rows()
        for cell in row
        if "Pass" in str(cell.value or "") and "Fail" in str(cell.value or "")
    ]
    assert len(pass_cells) == 30
    assert sum("■Pass" in str(cell.value) for cell in pass_cells) == 16
    assert all(
        "■Fail" not in str(cell.value) and "☑Fail" not in str(cell.value) for cell in pass_cells
    )
    na_cells = [cell for row in sheet.iter_rows() for cell in row if "N/A" in str(cell.value or "")]
    assert sum("■N/A" in str(cell.value) or "☑N/A" in str(cell.value) for cell in na_cells) == 9
    original_workbook = load_workbook(TEMPLATE)
    original = original_workbook["BT380"]
    assert isinstance(original, Worksheet)
    assert {str(item) for item in sheet.merged_cells.ranges} == {
        str(item) for item in original.merged_cells.ranges
    }
    assert len(sheet._images) == len(original._images)  # pyright: ignore[reportAttributeAccessIssue]
    assert sheet.print_area == original.print_area
    assert sheet.sheet_properties.pageSetUpPr == original.sheet_properties.pageSetUpPr
    assert sheet.page_setup.orientation == original.page_setup.orientation
    assert sheet.page_setup.paperSize == original.page_setup.paperSize
    for field in ("left", "right", "top", "bottom", "header", "footer"):
        assert getattr(sheet.page_margins, field) == pytest.approx(
            getattr(original.page_margins, field)
        )
    assert sheet.freeze_panes == original.freeze_panes
    assert sheet.print_title_rows == original.print_title_rows
    assert sheet.print_title_cols == original.print_title_cols
    assert len(sheet.data_validations.dataValidation) == len(
        original.data_validations.dataValidation
    )
    assert len(sheet.conditional_formatting) == len(original.conditional_formatting)
    assert {key: (item.height, item.hidden) for key, item in sheet.row_dimensions.items()} == {
        key: (item.height, item.hidden) for key, item in original.row_dimensions.items()
    }
    assert {key: (item.width, item.hidden) for key, item in sheet.column_dimensions.items()} == {
        key: (item.width, item.hidden) for key, item in original.column_dimensions.items()
    }
    assert SOURCE.read_bytes() and hashlib.sha256(SOURCE.read_bytes()).digest() == source_hash
    assert hashlib.sha256(TEMPLATE.read_bytes()).digest() == template_hash
    workbook.close()
    original_workbook.close()


def test_all_model_sheets_accept_the_audited_pass_signature() -> None:
    workbook = load_workbook(TEMPLATE, read_only=True)
    for model, expected in PASS_COUNTS.items():
        sheet = workbook[model]
        validation_cells = [
            cell
            for row in sheet.iter_rows()
            for cell in row
            if "Pass" in str(cell.value or "") and "Fail" in str(cell.value or "")
        ]
        assert len(validation_cells) == expected
    workbook.close()


def test_required_formula_without_cache_is_rejected(tmp_path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    assert isinstance(sheet, Worksheet)
    headers = [
        "서비스번호/의뢰일자",
        "접수자",
        "의뢰자",
        "병원명",
        "Model",
        "불량대분류",
        "증상/요청사항",
        "대응조치",
        "처리자",
        "처리완료일",
        "처리완료월",
    ]
    sheet.append(headers)
    sheet.append(
        [
            "DS26070601",
            "장진영",
            "국제메디칼",
            "미즈피아병원",
            "BT380",
            "화면불량",
            "=1+1",
            "교체",
            "장진영",
            "7/6",
            "7월",
        ]
    )
    path = tmp_path / "formula.xlsx"
    workbook.save(path)
    workbook.close()
    with pytest.raises(ValidationError, match="생성할 데이터 행"):
        read_source(path)


def test_unknown_model_creates_no_output(tmp_path: Path) -> None:
    record = replace(read_source(SOURCE).records[0], model="UNKNOWN")
    with pytest.raises(ValidationError, match="지원하지 않는 모델"):
        generate_workbook(TEMPLATE, tmp_path, record)
    assert not list(tmp_path.iterdir())


def test_bt350l_uses_bt350_template_and_preserves_original_model(tmp_path: Path) -> None:
    base = next(item for item in read_source(SOURCE).records if item.service_number == "DS26070601")
    record = replace(
        base,
        model="BT350L",
        defect_category="DOP 불량",
        service_details="도플러 감도 저하",
    )

    output = generate_workbook(TEMPLATE, tmp_path, record)
    workbook = load_workbook(output, data_only=True)
    sheet = workbook["BT350"]

    assert workbook.sheetnames == ["BT350"]
    assert sheet["O8"].value == "BT350L"
    assert output.name == "서비스검증결과서_DS26070601_국제메디칼(미즈피아병원).xlsx"
    assert "■Pass" in str(sheet["BO35"].value)
    assert "■N/A" in str(sheet["CA37"].value)
    workbook.close()


def test_all_pass_alias_model_does_not_require_conditional_group_mapping(
    tmp_path: Path,
) -> None:
    base = next(item for item in read_source(SOURCE).records if item.service_number == "DS26070601")
    record = replace(base, model="BT200L")

    output = generate_workbook(TEMPLATE, tmp_path, record)
    workbook = load_workbook(output, data_only=True)
    sheet = workbook["BT200"]
    pass_cells = [
        cell
        for row in sheet.iter_rows()
        for cell in row
        if "Pass" in str(cell.value or "") and "Fail" in str(cell.value or "")
    ]

    assert sheet["O8"].value == "BT200L"
    assert len(pass_cells) == PASS_COUNTS["BT200"]
    assert all("■Pass" in str(cell.value) for cell in pass_cells)
    workbook.close()


def test_bt740_links_ecg_spo2_nibp_and_marks_other_optional_groups_na(
    tmp_path: Path,
) -> None:
    base = next(item for item in read_source(SOURCE).records if item.service_number == "DS26070601")
    record = replace(
        base,
        model="BT740",
        defect_category="SpO2 불량",
        service_details="산소포화도 측정 안됨",
    )

    output = generate_workbook(TEMPLATE, tmp_path, record)
    workbook = load_workbook(output, data_only=True)
    sheet = workbook["BT740"]

    for row in (27, 28, 29, 30, 31):
        assert "■Pass" in str(sheet.cell(row, 67).value)
    for row in (32, 33, 34, 35, 36):
        assert "■N/A" in str(sheet.cell(row, 79).value)
    workbook.close()


def test_duplicate_required_header_is_rejected(tmp_path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    assert isinstance(sheet, Worksheet)
    sheet.append(
        [
            "서비스번호/의뢰일자",
            "서비스번호/의뢰일자",
            "접수자",
            "의뢰자",
            "병원명",
            "Model",
            "불량대분류",
            "증상/요청사항",
            "대응조치",
            "처리자",
            "처리완료일",
            "처리완료월",
        ]
    )
    path = tmp_path / "duplicate.xlsx"
    workbook.save(path)
    workbook.close()
    with pytest.raises(ValidationError, match="중복 헤더"):
        read_source(path)


def test_customer_config_supports_keywords_and_exact_overrides(tmp_path: Path) -> None:
    path = tmp_path / "customer_overrides.json"
    path.write_text(
        '{"company_keywords":["센터"],"overrides":{"홍길동":"대리점"}}',
        encoding="utf-8",
    )
    config = load_overrides(path)
    assert config.company_keywords == ("센터",)
    assert config.overrides["홍길동"].value == "대리점"


def test_multiple_matching_source_sheets_are_combined_and_invalid_rows_skipped(
    tmp_path: Path,
) -> None:
    from openpyxl import Workbook

    headers = [
        "서비스번호/의뢰일자",
        "접수자",
        "의뢰자",
        "병원명",
        "Model",
        "불량대분류",
        "증상/요청사항",
        "대응조치",
        "처리자",
        "처리완료일",
        "처리완료월",
    ]
    workbook = Workbook()
    first = workbook.active
    assert isinstance(first, Worksheet)
    first.append(headers)
    first.append(
        [
            "DS26070601",
            "장진영",
            "국제메디칼",
            "미즈피아병원",
            "BT380",
            "화면불량",
            "화면불량",
            "케이블 교체",
            "장진영",
            "7/6",
            "7월",
        ]
    )
    first.append(["DS26070602", "장진영", "", "", "BT380", "미완료"])
    second = workbook.create_sheet("두번째")
    second.append(headers)
    second.append(
        [
            "DS26070701",
            "장진영",
            "국제메디칼",
            "미즈피아병원",
            "BT380",
            "화면불량",
            "화면불량",
            "케이블 교체",
            "장진영",
            "7/7",
            "7월",
        ]
    )
    path = tmp_path / "multiple.xlsx"
    workbook.save(path)
    workbook.close()
    source = read_source(path)
    assert source.sheet_names == ("Sheet", "두번째")
    assert [record.service_number for record in source.records] == ["DS26070601", "DS26070701"]
    assert source.skipped_rows == 1
    assert len(source.rejected_rows) == 1
    assert source.rejected_rows[0].service_number == "DS26070602"
    assert "처리완료일" in source.rejected_rows[0].reason


def test_create_reports_generates_one_file_per_selected_record(tmp_path: Path) -> None:
    records = read_source(SOURCE).records
    first = next(record for record in records if record.service_number == "DS26070601")
    second = replace(first, service_number="DS26070602")
    result = create_reports(TEMPLATE, tmp_path, (first, second))
    assert len(result.outputs) == 2
    assert all(output.exists() for output in result.outputs)
    assert not result.failures


def test_create_reports_preserves_successes_when_later_record_fails(
    tmp_path: Path,
) -> None:
    records = read_source(SOURCE).records
    first = next(record for record in records if record.service_number == "DS26070601")
    invalid = replace(first, service_number="DS26070602", model="UNKNOWN")
    last = replace(first, service_number="DS26070603")

    result = create_reports(TEMPLATE, tmp_path, (first, invalid, last))

    assert [record.service_number for record in result.successful_records] == [
        "DS26070601",
        "DS26070603",
    ]
    assert len(result.outputs) == 2
    assert all(output.exists() for output in result.outputs)
    assert [failure.record.service_number for failure in result.failures] == [
        "DS26070602"
    ]
    assert "지원하지 않는 모델" in result.failures[0].detail
    append_issuance_log(tmp_path, result.successful_records, result.outputs)
    counts = load_issued_counts(tmp_path)
    assert counts[record_key(first)] == 1
    assert counts[record_key(last)] == 1


def test_create_reports_rejects_more_than_500_records(tmp_path: Path) -> None:
    record = read_source(SOURCE).records[0]

    with pytest.raises(ValidationError, match="최대 500개"):
        create_reports(TEMPLATE, tmp_path, (record,) * 501)

    assert not list(tmp_path.iterdir())


def test_create_reports_surfaces_worker_validation_error(tmp_path: Path) -> None:
    record = next(
        item for item in read_source(SOURCE).records if item.service_number == "DS26070601"
    )

    result = create_reports(tmp_path / "missing-template.xlsx", tmp_path, (record,))

    assert not result.outputs
    assert len(result.failures) == 1
    assert "검증결과서 양식 파일을 읽을 수 없습니다" in result.failures[0].detail


def test_issuance_history_and_status_report_distinguish_duplicate_rows(
    tmp_path: Path,
) -> None:
    record = next(
        item for item in read_source(SOURCE).records if item.service_number == "DS26070601"
    )
    output = tmp_path / "issued.xlsx"
    output.touch()
    history = append_issuance_log(tmp_path, (record,), (output,))

    counts = load_issued_counts(tmp_path)

    assert history.exists()
    assert counts[record_key(record)] == 1
    source = SourceRows(
        sheet_names=("원본",),
        records=(record, record),
        skipped_rows=1,
        rejected_rows=(RejectedRow("원본", 12, "DS26070699", "처리완료일이 없습니다."),),
    )
    report = write_status_report(tmp_path, source, counts)
    workbook = load_workbook(report, data_only=True)
    sheet = workbook["발행현황"]
    statuses = [sheet.cell(row, 1).value for row in range(2, sheet.max_row + 1)]
    assert statuses == ["발행완료", "미발행", "발행불가"]
    assert sheet.cell(4, 10).value == "처리완료일이 없습니다."
    workbook.close()

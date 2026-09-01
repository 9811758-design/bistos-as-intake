from collections.abc import Sequence
from pathlib import Path
from typing import Final, assert_never

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from receivables_reconciliation.tracker_models import DepositTask, EntryStatus

_HEADERS: Final = (
    "입력 상태",
    "입금일",
    "입금액",
    "입금자명",
    "금융기관",
    "메일 제목",
    "수신 시각",
    "확인 메모",
)
_WIDTHS: Final = (14, 13, 16, 20, 18, 40, 20, 42)


def export_tasks(path: Path, tasks: Sequence[DepositTask]) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    sheet = workbook.create_sheet("입금메일 정리", 0)
    if default_sheet is not None:
        workbook.remove(default_sheet)
    sheet.append(_HEADERS)
    for task in tasks:
        sheet.append(
            (
                _status_label(task.status),
                task.deposit_date,
                task.amount,
                task.depositor_name,
                task.bank_name,
                task.subject,
                task.received_at,
                task.note,
            )
        )
    _format_sheet(sheet)
    workbook.save(path)


def _format_sheet(sheet: Worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1D709B")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for index, width in enumerate(_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet["B"][1:]:
        cell.number_format = "yyyy-mm-dd"
    for cell in sheet["C"][1:]:
        cell.number_format = "#,##0\"원\""
    for cell in sheet["G"][1:]:
        cell.number_format = "yyyy-mm-dd hh:mm"


def _status_label(status: EntryStatus) -> str:
    match status:
        case EntryStatus.PENDING:
            return "미입력"
        case EntryStatus.COMPLETED:
            return "입력 완료"
        case unreachable:
            assert_never(unreachable)

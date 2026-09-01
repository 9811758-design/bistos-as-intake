from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Final

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from .domain import (
    CustomerClass,
    ServiceRecord,
    ValidationError,
    normalize_text,
    safe_filename,
    unique_output_path,
)
from .package_fidelity import restore_selected_sheet_drawing
from .validation_rules import Result, ValidationPlan, build_validation_plan, normalize_rule_text

ANCHORS: Final = {
    "service_number": "BO1",
    "receipt_date": "O4",
    "receiver": "BE4",
    "customer_class": "O5",
    "customer_name": "BE5",
    "model": "O8",
    "service_details": "O9",
    "processing_details": "O12",
    "completion_date": "O13",
    "processor": "BE13",
}
PASS_COUNTS: Final = {
    "BT36": 8,
    "BT100": 7,
    "BT200": 10,
    "BT220C": 14,
    "BT250": 15,
    "BT300": 19,
    "BT350": 25,
    "BT380": 30,
    "BT400": 7,
    "BT410": 9,
    "BT450": 9,
    "BT500": 26,
    "BT550": 22,
    "BT700": 18,
    "신규BT700": 10,
    "BT710": 11,
    "BT720": 14,
    "BT740": 20,
    "BT770": 20,
    "BT780": 20,
}
EXPECTED_MERGES: Final = {
    "BO1": "BO1:CF1",
    "O4": "O4:AP4",
    "BE4": "BE4:CF4",
    "O5": "O5:AP5",
    "BE5": "BE5:CF5",
    "O8": "O8:CF8",
    "O9": "O9:CF9",
    "O12": "O12:CF12",
    "O13": "O13:AP13",
    "BE13": "BE13:CF13",
}
EXPECTED_LABELS: Final = {
    "A4": ("접수일",),
    "A5": ("고객구분",),
    "A8": ("Model",),
    "A9": ("서비스 내용", "불만 내용"),
    "A12": ("처리 내역",),
    "A13": ("처리 완료일",),
    "A15": ("검증 결과",),
}
GROUP_ALIASES: Final = {
    "main": ("cpumodulebdmainbdassy", "fetalmonitorbdmainbdassy", "mainbd"),
    "lcd": ("lcdheadassy",),
    "lcd_touch": ("lcd및touch", "lcdtouch"),
    "led": ("ledassy",),
    "key": ("keyknobpowerswitchbdassy", "knobpowerswitchbdassy", "keyknob"),
    "printer": ("printengine", "printer"),
    "dop": ("dopprobe",),
    "uc": ("ucprobe",),
    "mark": ("mark",),
    "ast": ("ast",),
    "communication": ("유무선통신",),
    "spo2": ("spo2",),
    "ecg": ("ecg",),
    "nibp": ("nibp",),
    "temp": ("temp",),
    "co2": ("co2",),
    "ibp": ("ibp",),
    "cms": ("cms",),
}


@dataclass(frozen=True, slots=True)
class ChecklistGroup:
    group_id: str
    pass_cells: tuple[Cell, ...]
    na_cells: tuple[Cell, ...]


def _pass_cells(sheet: Worksheet) -> list[Cell]:
    return [
        cell
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell, Cell)
        and "Pass" in str(cell.value or "")
        and "Fail" in str(cell.value or "")
    ]


def _group_id(label: str) -> str | None:
    normalized = normalize_rule_text(label)
    for group_id, aliases in GROUP_ALIASES.items():
        if normalized in aliases:
            return group_id
    return None


def _checklist_groups(sheet: Worksheet) -> tuple[ChecklistGroup, ...]:
    grouped_pass: dict[str, list[Cell]] = {}
    grouped_na: dict[str, list[Cell]] = {}
    current: str | None = None
    for row in range(1, sheet.max_row + 1):
        label = str(sheet.cell(row, 1).value or "").strip()
        if label:
            current = _group_id(label)
        pass_cells = [
            cell
            for cell in sheet[row]
            if isinstance(cell, Cell)
            and "Pass" in str(cell.value or "")
            and "Fail" in str(cell.value or "")
        ]
        na_cells = [
            cell for cell in sheet[row] if isinstance(cell, Cell) and "N/A" in str(cell.value or "")
        ]
        if (pass_cells or na_cells) and current is None:
            raise ValidationError(f"양식의 검증 그룹을 인식할 수 없습니다: {row}행")
        if current is not None:
            grouped_pass.setdefault(current, []).extend(pass_cells)
            grouped_na.setdefault(current, []).extend(na_cells)
    return tuple(
        ChecklistGroup(group_id, tuple(cells), tuple(grouped_na.get(group_id, [])))
        for group_id, cells in grouped_pass.items()
    )


def _set_pass(cell: Cell, selected: bool) -> None:
    value = str(cell.value or "")
    for checked, unchecked in (("■Pass", "□Pass"), ("☑Pass", "☐Pass")):
        value = value.replace(checked, unchecked)
    if selected:
        value = value.replace("□Pass", "■Pass").replace("☐Pass", "☑Pass")
    cell.value = value


def _set_na(cell: Cell, selected: bool) -> None:
    value = str(cell.value or "")
    for checked, unchecked in (("■N/A", "□N/A"), ("☑N/A", "☐N/A")):
        value = value.replace(checked, unchecked)
    if selected:
        value = value.replace("□N/A", "■N/A").replace("☐N/A", "☑N/A")
    cell.value = value


def _apply_validation_plan(sheet: Worksheet, plan: ValidationPlan) -> None:
    if plan.all_pass:
        for cell in _pass_cells(sheet):
            _set_pass(cell, True)
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell, Cell) and "N/A" in str(cell.value or ""):
                    _set_na(cell, False)
        return
    groups = _checklist_groups(sheet)
    available = {group.group_id for group in groups}
    expected = {decision.group_id for decision in plan.decisions}
    if available != expected:
        missing = ", ".join(sorted(expected - available)) or "없음"
        unknown = ", ".join(sorted(available - expected)) or "없음"
        raise ValidationError(
            f"양식의 검증 그룹이 규칙과 다릅니다. 누락: {missing}, 미확인: {unknown}"
        )
    decisions = {decision.group_id: decision.result for decision in plan.decisions}
    for group in groups:
        result = decisions[group.group_id]
        for cell in group.pass_cells:
            _set_pass(cell, result is Result.PASS)
        for index, cell in enumerate(group.na_cells):
            _set_na(cell, result is Result.NA and index == 0)
        if result is Result.NA and not group.na_cells:
            raise ValidationError(f"양식에 N/A 선택란이 없습니다: {group.group_id}")


def _checked_customer_class(category: CustomerClass) -> str:
    return "      ".join(
        f"{'■' if category is item else '□'}{item.value}" for item in CustomerClass
    )


def _display_date(value: date) -> str:
    return f"{value.year:04d} 년 {value.month:02d} 월 {value.day:02d} 일"


def generate_workbook(template: Path, output_folder: Path, record: ServiceRecord) -> Path:
    plan = build_validation_plan(record)
    try:
        workbook = load_workbook(template)
    except Exception as exc:
        raise ValidationError("검증결과서 양식 파일을 읽을 수 없습니다.") from exc
    if plan.template_model not in workbook.sheetnames:
        workbook.close()
        raise ValidationError(f"양식에 모델 시트가 없습니다: {plan.template_model}")
    sheet = workbook[plan.template_model]
    merged = {str(item) for item in sheet.merged_cells.ranges}
    missing_merges = [expected for expected in EXPECTED_MERGES.values() if expected not in merged]
    if missing_merges:
        workbook.close()
        raise ValidationError(f"양식의 입력 영역이 변경되었습니다: {', '.join(missing_merges)}")
    changed_labels = [
        address
        for address, expected_labels in EXPECTED_LABELS.items()
        if normalize_text(str(sheet[address].value or ""))
        not in {normalize_text(label) for label in expected_labels}
    ]
    if changed_labels:
        workbook.close()
        raise ValidationError(f"양식의 항목명이 변경되었습니다: {', '.join(changed_labels)}")
    cells = _pass_cells(sheet)
    expected = PASS_COUNTS[plan.template_model]
    if len(cells) != expected:
        workbook.close()
        raise ValidationError(f"검증 항목 수가 예상과 다릅니다: {len(cells)} / {expected}")
    for other in tuple(workbook.worksheets):
        if other.title != plan.template_model:
            workbook.remove(other)
    values = {
        "service_number": record.service_number,
        "receipt_date": _display_date(record.receipt_date),
        "receiver": record.receiver,
        "customer_class": _checked_customer_class(record.customer.category),
        "customer_name": record.customer.display_name,
        "model": record.model,
        "service_details": record.service_details,
        "processing_details": record.processing_details,
        "completion_date": _display_date(record.completion_date),
        "processor": record.processor,
    }
    for key, anchor in ANCHORS.items():
        sheet[anchor] = values[key]
    processing_font = copy(sheet[ANCHORS["processing_details"]].font)
    processing_font.sz = 10
    sheet[ANCHORS["processing_details"]].font = processing_font
    _apply_validation_plan(sheet, plan)
    output_folder.mkdir(parents=True, exist_ok=True)
    output = unique_output_path(output_folder, safe_filename(record))
    try:
        workbook.save(output)
    finally:
        workbook.close()
    restore_selected_sheet_drawing(template, output, plan.template_model)
    return output

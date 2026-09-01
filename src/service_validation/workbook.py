from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Final

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .domain import CustomerClass, ServiceRecord, ValidationError, normalize_text
from .record_builder import RecordFields, build_service_record

HEADER_ALIASES: Final = {
    "service_number": ("서비스번호/의뢰일자",),
    "receiver": ("접수자",),
    "requester": ("의뢰자", "의뢰자/고객", "의뢰자(고객,영업담당자)"),
    "hospital": ("병원명",),
    "model": ("Model", "모델"),
    "defect_category": ("불량대분류",),
    "service_details": ("증상/요청사항",),
    "processing_details": ("대응조치",),
    "processor": ("처리자",),
    "completion_date": ("처리완료일",),
    "completion_month": ("처리완료월", "월"),
}


@dataclass(frozen=True, slots=True)
class SourceRows:
    sheet_names: tuple[str, ...]
    records: tuple[ServiceRecord, ...]
    skipped_rows: int
    rejected_rows: tuple[RejectedRow, ...] = ()


@dataclass(frozen=True, slots=True)
class RejectedRow:
    sheet_name: str
    row_number: int
    service_number: str
    reason: str


@dataclass(frozen=True, slots=True)
class CustomerConfig:
    overrides: dict[str, CustomerClass]
    company_keywords: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class SheetPair:
    formula: Worksheet
    values: Worksheet
    columns: dict[str, int]


def _header_map(sheet: Worksheet, row_number: int) -> dict[str, int] | None:
    normalized: dict[str, list[int]] = {}
    for cell in sheet[row_number]:
        key = normalize_text(str(cell.value or "")).lower()
        if key and isinstance(cell.column, int):
            normalized.setdefault(key, []).append(cell.column)
    result: dict[str, int] = {}
    for semantic, aliases in HEADER_ALIASES.items():
        matching = next(
            (
                columns
                for alias in aliases
                if (columns := normalized.get(normalize_text(alias).lower(), []))
            ),
            [],
        )
        if not matching:
            return None
        if len(matching) != 1:
            raise ValidationError(f"중복 헤더가 있습니다: {aliases[0]}")
        result[semantic] = matching[0]
    return result


def _find_source_sheets(workbook: Workbook) -> tuple[tuple[Worksheet, dict[str, int]], ...]:
    matches: list[tuple[Worksheet, dict[str, int]]] = []
    for sheet in workbook.worksheets:
        for row_number in range(1, min(sheet.max_row, 20) + 1):
            mapping = _header_map(sheet, row_number)
            if mapping is not None:
                mapping["header_row"] = row_number
                matches.append((sheet, mapping))
                break
    if not matches:
        raise ValidationError("필수 헤더가 있는 원본 시트를 찾을 수 없습니다.")
    return tuple(matches)


def load_overrides(path: Path) -> CustomerConfig:
    if not path.exists():
        return CustomerConfig({}, ())
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValidationError("고객 구분 설정 파일을 읽을 수 없습니다.") from exc
    if not isinstance(raw, dict):
        raise ValidationError("고객 구분 설정은 JSON 객체여야 합니다.")
    keywords_raw = raw.get("company_keywords", [])
    overrides_raw = raw.get("overrides", raw)
    if not isinstance(keywords_raw, list) or not all(
        isinstance(item, str) for item in keywords_raw
    ):
        raise ValidationError("company_keywords는 문자열 배열이어야 합니다.")
    if not isinstance(overrides_raw, dict):
        raise ValidationError("overrides는 JSON 객체여야 합니다.")
    result: dict[str, CustomerClass] = {}
    for key, value in overrides_raw.items():
        if key in {"company_keywords", "overrides"}:
            continue
        if isinstance(key, str) and isinstance(value, str):
            try:
                result[key] = CustomerClass(value)
            except ValueError as exc:
                raise ValidationError(f"알 수 없는 고객 구분: {value}") from exc
    return CustomerConfig(result, tuple(keywords_raw))


def read_source(path: Path, config: CustomerConfig | None = None) -> SourceRows:
    try:
        formula_book = load_workbook(path, data_only=False, read_only=False)
        value_book = load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:
        raise ValidationError("원본 엑셀 파일을 읽을 수 없습니다.") from exc
    try:
        matches = _find_source_sheets(formula_book)
        records: list[ServiceRecord] = []
        rejected_rows: list[RejectedRow] = []
        for formula_sheet, columns in matches:
            pair = SheetPair(formula_sheet, value_book[formula_sheet.title], columns)
            sheet_records, sheet_rejected = _read_records(pair, config)
            records.extend(sheet_records)
            rejected_rows.extend(sheet_rejected)
    finally:
        formula_book.close()
        value_book.close()
    if not records:
        raise ValidationError("생성할 데이터 행이 없습니다.")
    return SourceRows(
        sheet_names=tuple(sheet.title for sheet, _ in matches),
        records=tuple(records),
        skipped_rows=len(rejected_rows),
        rejected_rows=tuple(rejected_rows),
    )


def _read_records(
    pair: SheetPair,
    config: CustomerConfig | None,
) -> tuple[list[ServiceRecord], list[RejectedRow]]:
    records: list[ServiceRecord] = []
    rejected_rows: list[RejectedRow] = []
    for row in range(pair.columns["header_row"] + 1, pair.formula.max_row + 1):
        try:
            record = _read_record(pair, row, config)
        except ValidationError as exc:
            raw_service_number = pair.formula.cell(row, pair.columns["service_number"]).value
            rejected_rows.append(
                RejectedRow(
                    sheet_name=pair.formula.title,
                    row_number=row,
                    service_number=str(raw_service_number or "").strip(),
                    reason=str(exc),
                )
            )
            continue
        if record is not None:
            records.append(record)
    return records, rejected_rows


def _read_record(pair: SheetPair, row: int, config: CustomerConfig | None) -> ServiceRecord | None:
    values: dict[str, object] = {}
    for key, column in pair.columns.items():
        if key == "header_row":
            continue
        formula_value = pair.formula.cell(row, column).value
        cached_value = pair.values.cell(row, column).value
        if isinstance(formula_value, str) and formula_value.startswith("="):
            if cached_value in (None, ""):
                raise ValidationError(f"{row}행 {key}: 수식의 저장된 계산값이 없습니다.")
            values[key] = cached_value
        else:
            values[key] = formula_value
    service_number = str(values["service_number"] or "").strip()
    if not service_number:
        return None
    completion = values["completion_date"]
    if completion in (None, ""):
        raise ValidationError(f"{service_number}: 처리완료일이 없습니다.")
    raw_completion = (
        completion
        if isinstance(completion, (str, int, float, date, datetime))
        else str(completion)
    )
    return build_service_record(
        RecordFields(
        service_number=service_number,
        receiver=str(values["receiver"] or "").strip(),
        requester=str(values["requester"] or "").strip(),
        hospital=str(values["hospital"] or "").strip(),
        model=str(values["model"] or "").strip(),
        defect_category=str(values["defect_category"] or "").strip(),
        service_details=str(values["service_details"] or "").strip(),
        processing_details=str(values["processing_details"] or "").strip(),
        completion_date=raw_completion,
        completion_month=str(values["completion_month"] or "").strip(),
        processor=str(values["processor"] or "").strip(),
        ),
        config,
    )

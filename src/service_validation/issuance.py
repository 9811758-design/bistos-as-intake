from __future__ import annotations

import hashlib
import json
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Final

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .domain import ServiceRecord, ValidationError, unique_output_path
from .validation_rules import build_validation_plan
from .workbook import SourceRows

LOG_FILENAME: Final = "서비스검증결과서_발행이력.xlsx"
LOG_HEADERS: Final = (
    "발행일시",
    "상태",
    "서비스번호",
    "의뢰자",
    "병원명",
    "모델",
    "증상/요청사항",
    "처리완료일",
    "결과파일",
    "고유키",
    "불량대분류",
    "검증판정",
)
STATUS_HEADERS: Final = (
    "상태",
    "서비스번호",
    "의뢰자",
    "병원명",
    "모델",
    "불량대분류",
    "증상/요청사항",
    "대응조치",
    "처리완료일",
    "발행불가 사유",
    "원본시트",
    "원본행",
    "검증판정",
)


def record_key(record: ServiceRecord) -> str:
    values = (
        record.service_number,
        record.requester,
        record.hospital,
        record.model,
        record.service_details,
        record.processing_details,
        record.completion_date.isoformat(),
    )
    serialized = json.dumps(values, ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()


def issuance_log_path(output_folder: Path) -> Path:
    return output_folder / LOG_FILENAME


def load_issued_counts(output_folder: Path) -> Counter[str]:
    path = issuance_log_path(output_folder)
    if not path.exists():
        return Counter()
    try:
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:
        raise ValidationError("기존 발행이력 파일을 읽을 수 없습니다.") from exc
    try:
        sheet = workbook.active
        if not isinstance(sheet, Worksheet):
            raise ValidationError("발행이력 워크시트를 찾을 수 없습니다.")
        headers = [str(cell.value or "") for cell in sheet[1]]
        if "고유키" not in headers:
            raise ValidationError("발행이력 파일의 형식이 올바르지 않습니다.")
        key_column = headers.index("고유키") + 1
        return Counter(
            str(sheet.cell(row, key_column).value)
            for row in range(2, sheet.max_row + 1)
            if sheet.cell(row, key_column).value
        )
    finally:
        workbook.close()


def append_issuance_log(
    output_folder: Path,
    records: tuple[ServiceRecord, ...],
    outputs: tuple[Path, ...],
) -> Path:
    if len(records) != len(outputs):
        raise ValidationError("발행 결과와 선택 데이터의 개수가 일치하지 않습니다.")
    output_folder.mkdir(parents=True, exist_ok=True)
    path = issuance_log_path(output_folder)
    if path.exists():
        try:
            workbook = load_workbook(path)
        except Exception as exc:
            raise ValidationError("기존 발행이력 파일을 열 수 없습니다.") from exc
        sheet = workbook.active
        if not isinstance(sheet, Worksheet):
            workbook.close()
            raise ValidationError("발행이력 워크시트를 찾을 수 없습니다.")
    else:
        workbook = Workbook()
        sheet = workbook.active
        if not isinstance(sheet, Worksheet):
            workbook.close()
            raise ValidationError("발행이력 워크시트를 만들 수 없습니다.")
        sheet.title = "발행이력"
        sheet.append(LOG_HEADERS)
        _style_sheet(sheet, len(LOG_HEADERS))
        sheet.column_dimensions["J"].hidden = True
    headers = [str(cell.value or "") for cell in sheet[1]]
    for header in LOG_HEADERS:
        if header not in headers:
            sheet.cell(1, len(headers) + 1, header)
            headers.append(header)
    _style_sheet(sheet, len(headers))
    key_column = headers.index("고유키") + 1
    sheet.column_dimensions[get_column_letter(key_column)].hidden = True
    issued_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for record, output in zip(records, outputs, strict=True):
        values = {
            "발행일시": issued_at,
            "상태": "발행완료",
            "서비스번호": record.service_number,
            "의뢰자": record.requester,
            "병원명": record.hospital,
            "모델": record.model,
            "증상/요청사항": record.service_details,
            "처리완료일": record.completion_date.isoformat(),
            "결과파일": str(output),
            "고유키": record_key(record),
            "불량대분류": record.defect_category,
            "검증판정": build_validation_plan(record).summary,
        }
        sheet.append(tuple(values.get(header, "") for header in headers))
    try:
        workbook.save(path)
    except PermissionError as exc:
        raise ValidationError(
            "발행이력 파일이 열려 있습니다. 파일을 닫고 다시 시도하세요."
        ) from exc
    finally:
        workbook.close()
    return path


def write_status_report(
    output_folder: Path,
    source: SourceRows,
    issued_counts: Counter[str],
) -> Path:
    output_folder.mkdir(parents=True, exist_ok=True)
    filename = f"서비스검증결과서_발행현황_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    path = unique_output_path(output_folder, filename)
    workbook = Workbook()
    sheet = workbook.active
    if not isinstance(sheet, Worksheet):
        workbook.close()
        raise ValidationError("발행현황 워크시트를 만들 수 없습니다.")
    sheet.title = "발행현황"
    sheet.append(STATUS_HEADERS)
    remaining = issued_counts.copy()
    status_counts: Counter[str] = Counter()
    for record in source.records:
        key = record_key(record)
        issued = remaining[key] > 0
        if issued:
            remaining[key] -= 1
        status = "발행완료" if issued else "미발행"
        status_counts[status] += 1
        sheet.append(
            (
                status,
                record.service_number,
                record.requester,
                record.hospital,
                record.model,
                record.defect_category,
                record.service_details,
                record.processing_details,
                record.completion_date.isoformat(),
                "",
                "",
                "",
                build_validation_plan(record).summary,
            )
        )
    for rejected in source.rejected_rows:
        status_counts["발행불가"] += 1
        sheet.append(
            (
                "발행불가",
                rejected.service_number,
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                rejected.reason,
                rejected.sheet_name,
                rejected.row_number,
                "",
            )
        )
    _style_sheet(sheet, len(STATUS_HEADERS))
    widths = (13, 18, 18, 22, 12, 18, 38, 38, 15, 48, 28, 10, 70)
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    summary = workbook.create_sheet("요약", 0)
    summary.append(("항목", "건수"))
    summary.append(("전체 대상", sum(status_counts.values())))
    summary.append(("발행완료", status_counts["발행완료"]))
    summary.append(("미발행", status_counts["미발행"]))
    summary.append(("발행불가", status_counts["발행불가"]))
    summary.append(("현황 생성일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    summary.append(("원본시트", ", ".join(source.sheet_names)))
    _style_sheet(summary, 2)
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 45
    try:
        workbook.save(path)
    finally:
        workbook.close()
    return path


def _style_sheet(sheet: Worksheet, column_count: int) -> None:
    fill = PatternFill("solid", fgColor="172554")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(column_count)}1"
    sheet.row_dimensions[1].height = 24
